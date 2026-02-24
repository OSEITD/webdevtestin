from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import os

# Ensure docs dir exists
os.makedirs('docs', exist_ok=True)

# Test cases data
outlet_cases = [
    ("OUT-001","Login as Outlet user","1. Open outlet-app/index.php or login.php\n2. Enter valid outlet email and password\n3. Click Sign In","User is logged in and outlet_dashboard.php appears with outlet name","") ,
    ("OUT-002","Dashboard shows summary stats","1. Login as outlet user\n2. Open Dashboard","Cards show numbers for Total Parcels, Pending, At Outlet, In Transit","") ,
    ("OUT-003","Parcel Pool — apply filters","1. Go to pages/parcelpool.php\n2. Select Status = Pending, Relation = Incoming, Date = This Month\n3. Click Apply Filters","Table shows only parcels that match filters","") ,
    ("OUT-004","Parcel Pool — search","1. Open Parcel Pool\n2. Enter a track number or sender name in Search\n3. Press Enter or click Apply","Only matching parcels are shown; message if none found","") ,
    ("OUT-005","Parcel Pool — view details modal","1. In Parcel Pool click the eye icon for any row\n2. Wait for modal","Modal opens showing sender, receiver, weight, fee, route, status","") ,
    ("OUT-006","Parcel Pool — empty result message","1. Set filters that match nothing\n2. Click Apply","Table shows 'No parcels found matching your criteria'","") ,
    ("OUT-007","Parcel Registration — add new parcel","1. Go to pages/parcel_registration.php\n2. Click Add Parcel\n3. Fill valid sender/receiver and parcel fields\n4. Click Save","New parcel created, success message shown, appears in Parcel Pool","") ,
    ("OUT-008","Parcel Registration — validation errors","1. Open Add Parcel form\n2. Leave required fields empty\n3. Click Save","Form shows validation messages and prevents save","") ,
    ("OUT-009","Parcel Management — edit parcel","1. Open pages/parcel_management.php\n2. Open a parcel and change receiver or weight\n3. Save","Changes are saved and shown in list","") ,
    ("OUT-010","Parcel Management — cancel or delete parcel","1. Select a parcel\n2. Click Cancel or Delete\n3. Confirm","Parcel status shows cancelled or item is removed (with confirmation)","") ,
    ("OUT-011","Bulk assign parcels to trip","1. In Parcel Pool select multiple rows\n2. Click Assign to Trip\n3. Choose trip/vehicle and confirm","All selected parcels show status assigned and linked to trip","") ,
    ("OUT-012","Barcode / label generation","1. Open a parcel row\n2. Click Print Label or Generate Barcode","Barcode/label opens or downloads and matches tracking number","") ,
    ("OUT-013","Create trip with Trip Wizard","1. Go to pages/trip_wizard.php\n2. Click Create Trip\n3. Add stops/parcels and assign vehicle\n4. Save","Trip is created and visible in trips.php or manager_trips.php","") ,
    ("OUT-014","Manager assign driver to trip","1. Open manager_trips.php\n2. Select a trip\n3. Assign a driver and save","Trip shows assigned driver and status updates","") ,
    ("OUT-015","Start and complete trip","1. Open a trip in trips.php\n2. Click Start Trip\n3. Mark stops as Arrived/Completed","Trip status changes and timestamps set","") ,
    ("OUT-016","Assignment tracking (real-time)","1. Open pages/assignment_tracking.php\n2. Watch an assigned trip update in real time","Trip/assignment updates appear live on map/list","") ,
    ("OUT-017","View/export revenue report","1. Go to pages/revenue_report.php\n2. Select date range\n3. Click View and Export","Report shows totals and export downloads (CSV/PDF)","") ,
    ("OUT-018","Business customers — add & search","1. Open pages/business_customers.php\n2. Click Add and fill details\n3. Save\n4. Use Search","New business customer is created and visible via search","") ,
    ("OUT-019","Notifications — in-app and mark read","1. Trigger a parcel update or system alert\n2. Click pages/notifications.php\n3. Mark as read","Notification appears and changes to read state","") ,
    ("OUT-020","Outlet settings — update profile","1. Go to pages/outlet_settings.php\n2. Change outlet name, address or phone\n3. Save","Changes are saved and displayed on profile","") ,
    ("OUT-021","Change password","1. Go to Change Password\n2. Enter current and new password\n3. Save","Password updated; new password works at next login","") ,
    ("OUT-022","Help / Manage Help content","1. Open pages/manage_help.php\n2. Edit or add a help article\n3. Save\n4. Search via pages/help.php","Help article is saved and searchable","") ,
    ("OUT-023","Clear cache (admin utility)","1. Go to pages/clear_cache.php\n2. Click Clear Cache","Cache clears and success message shows; pages reflect fresh data","") ,
    ("OUT-024","Push subscription saved for manager","1. Trigger push subscribe flow in manager UI\n2. Allow browser notifications","Subscription is saved and notifications can be sent","") ,
    ("OUT-025","Table pagination & sorting","1. Open long parcel list\n2. Click next page and sort a column","Page navigation and sort update the rows correctly","") ,
    ("OUT-026","Access control — outlet pages protected","1. Open an outlet page while logged out\n2. Try direct URL access","User is redirected to login.php or shown access denied","") ,
    ("OUT-027","Mobile responsiveness — parcel pool","1. Open pages/parcelpool.php on a small screen\n2. View rows and actions","Table stacks to readable cards; actions remain usable","") ,
    ("OUT-028","Logout","1. Click Logout","User is logged out and returned to login page","")
]

driver_cases = [
    ("DRV-001","Driver login","1. Open driver login\n2. Enter driver credentials\n3. Click Sign In","Driver dashboard loads with driver name and assignments","") ,
    ("DRV-002","View assigned trips","1. Login as driver\n2. Open My Trips or dashboard","Assigned trips list is visible with statuses and next stops","") ,
    ("DRV-003","Scan parcel barcode (happy)","1. Open drivers/pages/scanner.php\n2. Scan barcode or enter tracking number\n3. Confirm","Parcel details open and scan is recorded","") ,
    ("DRV-004","Scan invalid barcode (negative)","1. Scan random/invalid code\n2. Try to confirm","App shows error 'Parcel not found' and prevents action","") ,
    ("DRV-005","Mark pickup at stop","1. Open a trip stop\n2. Click Picked Up","Parcel status updates and timestamp recorded","") ,
    ("DRV-006","Mark delivered + upload proof","1. On delivery click Mark Delivered\n2. Upload photo/signature\n3. Save","Parcel shows status delivered and proof is attached","") ,
    ("DRV-007","Proof upload invalid file type (negative)","1. Try to upload unsupported file (e.g., .exe)\n2. Save","App rejects file and shows validation message","") ,
    ("DRV-008","View route & navigation","1. Open drivers/pages/route.php for a trip\n2. Click View Route","Map shows route, stops, and estimated times","") ,
    ("DRV-009","Live tracking updates","1. Open drivers/pages/live-tracking.php\n2. Move device (or simulate)","Driver location updates on map and server receives location","") ,
    ("DRV-010","Uber-style live-tracking view","1. Open live-tracking-uber-style.php\n2. Observe","Alternate live-tracking UI displays and updates correctly","") ,
    ("DRV-011","Delivery history & filters","1. Open drivers/pages/delivery-history.php\n2. Filter by date/status","History lists matching deliveries and filters work","") ,
    ("DRV-012","Trip stop status flow (arrive/complete)","1. Start a trip\n2. Mark stop as Arrived then Complete","Stop status updates and trip progress advances","") ,
    ("DRV-013","Accept/decline assignment","1. Receive an assignment notification\n2. Click Accept or Decline","Assignment is accepted or returned for reassignment","") ,
    ("DRV-014","Update profile details","1. Go to profile page\n2. Change phone/photo\n3. Save","Profile updates persist and show in UI","") ,
    ("DRV-015","Offline mode then sync","1. Turn off network\n2. Mark a delivery completed\n3. Reconnect","Offline action is saved locally and syncs to server","") ,
    ("DRV-016","Receive push notification and act","1. Allow push for driver\n2. Trigger a push (new assignment)\n3. Tap notification","Driver sees notification and tapping opens correct screen","") ,
    ("DRV-017","Change availability (online/offline)","1. Toggle availability in dashboard\n2. Save","Driver status updates and dispatch sees correct availability","") ,
    ("DRV-018","GPS permission denied handling","1. Deny GPS permission when prompted\n2. Open live tracking","App shows clear message and fallback (manual update)","") ,
    ("DRV-019","Access control — driver pages","1. Try to open driver pages when logged out\n2. Direct URL access","Redirect to login or access denied shown","") ,
    ("DRV-020","Logout","1. Click Logout","Driver is logged out and returned to login screen","")
]

customer_cases = [
    ("CUS-001","Track parcel by tracking number","1. Open customer-app/track_parcel.php\n2. Enter a tracking number\n3. Click Track","Parcel summary and current status appear with timeline","") ,
    ("CUS-002","Track with invalid number (negative)","1. Enter wrong tracking number\n2. Click Track","Friendly error 'No parcel found' and suggestions shown","") ,
    ("CUS-003","Secure tracking with token/OTP","1. Open secure_tracking.html\n2. Enter tracking number and OTP/token\n3. Click Submit","User is taken to track_details.php with secure details","") ,
    ("CUS-004","Secure tracking — expired/invalid token","1. Enter expired or wrong token\n2. Submit","Page shows message 'Invalid or expired token' and no details","") ,
    ("CUS-005","View timeline & live map","1. On track_details.php click Show Timeline or Live Map\n2. Observe","Timeline shows events; map shows location if available","") ,
    ("CUS-006","View payment info & download receipt","1. Open Track Details for paid parcel\n2. View Payment section\n3. Click Download Receipt","Payment details shown and receipt downloads/opens","") ,
    ("CUS-007","Request redelivery / schedule pickup","1. On Track Details click Request Redelivery\n2. Choose date/time\n3. Submit","System confirms request and shows scheduled info","") ,
    ("CUS-008","Enable push notifications","1. On tracking page click Enable Notifications\n2. Allow browser push","Subscription saved and future updates show as push notifications","") ,
    ("CUS-009","Push notification opens tracking page","1. Receive a push update for parcel\n2. Tap notification","App/browser opens track_parcel.php?track=... or track_details.php","") ,
    ("CUS-010","Register account","1. Open Register page\n2. Enter name, email, phone, password\n3. Submit","Account created and user can login or is auto-logged in","") ,
    ("CUS-011","Forgot / Reset password","1. Click Forgot Password\n2. Enter email\n3. Follow reset link and set new password","Password reset succeeds and new password logs in","") ,
    ("CUS-012","Contact support / raise enquiry","1. On Track Details click Contact Support\n2. Enter message\n3. Submit","Support message is saved and confirmation shown","") ,
    ("CUS-013","GPS tracking (customer map)","1. Open gps_tracking.html and allow location\n2. Track parcel or view live map","Map shows parcel and/or your device location as expected","") ,
    ("CUS-014","Secure session expiry handling","1. Stay idle on Track Details until session expires\n2. Try to interact","App prompts for re-auth or shows session-expired message","") ,
    ("CUS-015","Logout","1. Click Logout (if logged in)","User is logged out and returned to public tracking page","")
]

# Create workbook and write sheets
wb = Workbook()
modules = [("Outlet", outlet_cases), ("Driver", driver_cases), ("Customer", customer_cases)]

for idx, (title, cases) in enumerate(modules):
    if idx == 0:
        ws = wb.active
        ws.title = title
    else:
        ws = wb.create_sheet(title=title)

    headers = ["Test Case ID", "Test Description", "Test Steps", "Expected Results", "Test Case Passed"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for case in cases:
        ws.append(list(case))

    widths = [18, 40, 60, 60, 18]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
        if col_idx in (3, 4):
            for cell in ws[get_column_letter(col_idx)]:
                cell.alignment = Alignment(wrapText=True, vertical='top')

path = os.path.join('docs', 'UAT_Test_Suites.xlsx')
wb.save(path)
print(f"WROTE: {path}")